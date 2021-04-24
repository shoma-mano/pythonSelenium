import openpyxl
from openpyxl.drawing.image import Image
import datetime
from common_module import postslackreport

#日付取得
dt_now=datetime.datetime.now()
month=str(dt_now.month)
day=str(dt_now.day)
hour=str(dt_now.hour)
minute=str(dt_now.minute)
if len(minute)==1:
    minute="0"+minute

def excute(data):
    # ワークブック読み込み
    book = openpyxl.load_workbook("excel//ResultTemplate.xlsx")

    # シートを取得し名前を決定
    ws = book.worksheets[0]
    ws.title = "実施結果"
    
    #実施日時タイトル書き込み
    title=ws["A2"]
    title.value=month+"月"+day+"日 "+hour+":"+minute+" 自動テスト実施結果"

    #書き込む行のセル取得
    cells = ws['B7':'E10']

    for index,row in enumerate(cells):
        for ind,cell in enumerate(row):
            try:
              cell.value=data.result[(ws[cell.row][0]).value][(ws[5][cell.column-1]).value]
            except:
                pass


    book.save("excel\\"+"実施結果"+'.xlsx')
    print("エクセルを作成しました")

    postslackreport.SendToSlack("実施結果")

